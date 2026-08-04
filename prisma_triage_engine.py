"""Phase 1 Prisma/Twistlock findings triage.

The module finds repeated vulnerability or policy patterns across distinct
resources and highlights package/image automation candidates. It never changes
cloud resources, executes scripts, or generates infrastructure code.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Optional, Union

import yaml


DEFAULT_CONFIG = {
    "recurring_min_resources": 2,
    "max_candidates_in_report": 25,
    "max_patterns_in_report": 50,
    "sheet_name": None,
    "output": {
        "report_path": "triage_report.md",
        "categorized_path": "categorized_findings.csv",
    },
    "column_aliases": {
        "finding_id": ["cve id", "policy", "policy name", "finding type", "alert name"],
        "resource_id": ["resource id", "resource", "resource name", "asset id"],
        "resource_name": ["hostname", "host name", "resource name"],
        "resource_type": ["type", "resource type", "asset type", "cloud type"],
        "severity": ["severity", "risk", "priority"],
        "component": ["source package", "packages", "package", "component"],
        "installed_version": ["package version", "installed version", "current version"],
        "fix_status": ["fix status", "remediation", "recommendation", "resolution"],
        "distro": ["distro", "distribution", "operating system"],
        "account": ["account id", "account", "account name", "cloud account"],
        "region": ["region", "cloud region"],
        "cluster": ["cluster", "cluster name"],
        "description": ["description", "finding description", "issue"],
        "owner": ["owner", "resource owner", "team"],
        "terraform_status": [
            "terraform status",
            "terraform managed",
            "iac status",
            "management status",
        ],
    },
}

SEVERITY_WEIGHT = {
    "critical": 5,
    "high": 4,
    "medium": 3,
    "low": 2,
    "informational": 1,
    "info": 1,
}


def load_config(config_path: Union[str, Path] = "config.yaml") -> dict[str, Any]:
    """Load YAML configuration and merge it over safe defaults."""
    config = {
        **DEFAULT_CONFIG,
        "output": {**DEFAULT_CONFIG["output"]},
        "column_aliases": {
            key: list(values) for key, values in DEFAULT_CONFIG["column_aliases"].items()
        },
    }
    path = Path(config_path)
    if not path.exists():
        return config
    loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    for key, value in loaded.items():
        if key in {"output", "column_aliases"} and isinstance(value, dict):
            config[key].update(value)
        else:
            config[key] = value
    return config


def _normalized(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_xlsx(path: Path, sheet_name: Optional[str]) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("XLSX input requires openpyxl from requirements.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [_clean(value) for value in next(rows)]
    except StopIteration:
        return []
    return [dict(zip(headers, row)) for row in rows]


def read_findings(
    path: Union[str, Path], sheet_name: Optional[str] = None
) -> list[dict[str, Any]]:
    """Read findings from CSV or XLSX."""
    input_path = Path(path)
    if input_path.suffix.lower() == ".csv":
        return _read_csv(input_path)
    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx(input_path, sheet_name)
    raise ValueError("Unsupported input format. Use CSV or XLSX.")


def normalize_findings(
    rows: Iterable[dict[str, Any]], column_aliases: dict[str, list[str]]
) -> list[dict[str, str]]:
    """Map report-specific headers to canonical triage fields."""
    alias_map = {
        field: [_normalized(alias) for alias in aliases + [field]]
        for field, aliases in column_aliases.items()
    }
    normalized_rows = []
    for source_row in rows:
        source = {_normalized(key): _clean(value) for key, value in source_row.items()}
        item = {}
        for field, aliases in alias_map.items():
            item[field] = next((source[alias] for alias in aliases if source.get(alias)), "")
        normalized_rows.append(item)
    return normalized_rows


def _effective_resource_id(finding: dict[str, str]) -> str:
    return finding["resource_id"] or finding["resource_name"]


def _pattern_key(finding: dict[str, str]) -> Optional[tuple[str, ...]]:
    finding_id = _normalized(finding["finding_id"])
    resource_type = _normalized(finding["resource_type"])
    if not finding_id or not resource_type or not _effective_resource_id(finding):
        return None
    return (
        finding_id,
        _normalized(finding["component"]),
        _normalized(finding["installed_version"]),
        _normalized(finding["fix_status"]),
        resource_type,
    )


def _reported_management_status(values: Iterable[str]) -> str:
    normalized = {_normalized(value) for value in values if value}
    managed = {"yes", "true", "managed", "terraform", "terraform managed"}
    unmanaged = {"no", "false", "unmanaged", "manual", "manually managed"}
    if normalized and normalized <= managed:
        return "Reported as Terraform-managed — validation required"
    if normalized and normalized <= unmanaged:
        return "Reported as unmanaged — validation required"
    if normalized:
        return "Mixed or unrecognized — validation required"
    return "Unknown — ownership validation required"


def _highest_severity(findings: Iterable[dict[str, str]]) -> str:
    return max(
        (finding["severity"] or "Unknown" for finding in findings),
        key=lambda value: SEVERITY_WEIGHT.get(value.lower(), 0),
        default="Unknown",
    )


def _route(category: str, management_status: str, has_component: bool) -> str:
    if category == "Needs More Information":
        return "Collect missing finding/resource context before routing"
    if category == "Unique":
        return "Manual review with the resource owner"
    if has_component:
        return "Package/image update candidate after ownership, compatibility, and rollout validation"
    if management_status.startswith("Reported as Terraform-managed"):
        return "Terraform template candidate after code/state ownership validation"
    if management_status.startswith("Reported as unmanaged"):
        return "Script-assisted or import candidate after human review"
    return "Validate ownership, then assess template, preventive control, or script-assisted treatment"


def _build_automation_candidates(patterns: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Roll repeated vulnerability patterns into package/image pain points."""
    grouped: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for pattern in patterns:
        if pattern["category"] != "Recurring" or not pattern["component"]:
            continue
        key = (
            _normalized(pattern["component"]),
            _normalized(pattern["installed_version"]),
            _normalized(pattern["distro"]),
            _normalized(pattern["resource_type"]),
        )
        grouped[key].append(pattern)

    candidates = []
    for members in grouped.values():
        resources = sorted({resource for member in members for resource in member["resources"]})
        finding_ids = sorted({member["finding_id"] for member in members})
        fix_statuses = sorted(
            {member["fix_status"] for member in members if member["fix_status"]}
        )
        severity = max(
            (member["severity"] for member in members),
            key=lambda value: SEVERITY_WEIGHT.get(value.lower(), 0),
        )
        candidates.append(
            {
                "component": members[0]["component"],
                "installed_version": members[0]["installed_version"],
                "distro": members[0]["distro"],
                "resource_type": members[0]["resource_type"],
                "severity": severity,
                "finding_count": len(finding_ids),
                "resource_count": len(resources),
                "fix_status_count": len(fix_statuses),
                "finding_ids": finding_ids,
                "resources": resources,
                "suggested_route": (
                    "Assess one controlled package/image update; validate the supported target "
                    "version, ownership, compatibility, rollout, rollback, and approval"
                ),
                "priority_score": (
                    SEVERITY_WEIGHT.get(severity.lower(), 0) * 100000
                    + len(finding_ids) * 100
                    + len(resources)
                ),
            }
        )
    return sorted(candidates, key=lambda item: item["priority_score"], reverse=True)


def classify_findings(
    findings: list[dict[str, str]], recurring_min_resources: int = 2
) -> dict[str, Any]:
    """Classify patterns by distinct affected resources, not duplicate row count."""
    if recurring_min_resources < 2:
        raise ValueError("recurring_min_resources must be at least 2")

    groups: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    incomplete = []
    for finding in findings:
        key = _pattern_key(finding)
        if key is None:
            incomplete.append(finding)
        else:
            groups[key].append(finding)

    patterns = []
    categorized_rows = []
    for members in groups.values():
        resources = sorted({_effective_resource_id(member) for member in members})
        category = "Recurring" if len(resources) >= recurring_min_resources else "Unique"
        management_status = _reported_management_status(
            member["terraform_status"] for member in members
        )
        severity = _highest_severity(members)
        pattern = {
            "category": category,
            "finding_id": members[0]["finding_id"],
            "resource_type": members[0]["resource_type"],
            "component": members[0]["component"],
            "installed_version": members[0]["installed_version"],
            "fix_status": members[0]["fix_status"],
            "distro": members[0]["distro"],
            "severity": severity,
            "row_count": len(members),
            "resource_count": len(resources),
            "resources": resources,
            "management_status": management_status,
            "suggested_route": _route(category, management_status, bool(members[0]["component"])),
            "priority_score": SEVERITY_WEIGHT.get(severity.lower(), 0) * 10000 + len(resources),
        }
        patterns.append(pattern)
        for member in members:
            categorized_rows.append(
                {**member, "category": category, "suggested_route": pattern["suggested_route"]}
            )

    for finding in incomplete:
        categorized_rows.append(
            {
                **finding,
                "category": "Needs More Information",
                "suggested_route": _route("Needs More Information", "Unknown", False),
            }
        )

    patterns.sort(key=lambda item: item["priority_score"], reverse=True)
    counts = Counter(row["category"] for row in categorized_rows)
    pattern_counts = Counter(pattern["category"] for pattern in patterns)
    return {
        "total": len(findings),
        "counts": dict(counts),
        "pattern_counts": dict(pattern_counts),
        "patterns": patterns,
        "automation_candidates": _build_automation_candidates(patterns),
        "categorized_rows": categorized_rows,
    }


def _examples(values: list[str], limit: int = 5) -> str:
    displayed = ", ".join(f"`{value}`" for value in values[:limit])
    if len(values) > limit:
        displayed += f" and {len(values) - limit} more"
    return displayed or "Not provided"


def render_report(
    result: dict[str, Any], max_candidates: int = 25, max_patterns: int = 50
) -> str:
    """Render a review-oriented Markdown report."""
    counts = result["counts"]
    pattern_counts = result["pattern_counts"]
    lines = [
        "# Prisma Findings Triage Report",
        "",
        "> Classification draft only. No cloud resources, scripts, images, packages, or Terraform changes were executed.",
        "",
        "## Executive Summary",
        "",
        f"- Total finding rows: **{result['total']}**",
        f"- Recurring rows: **{counts.get('Recurring', 0)}** across **{pattern_counts.get('Recurring', 0)}** patterns",
        f"- Unique rows: **{counts.get('Unique', 0)}** across **{pattern_counts.get('Unique', 0)}** patterns",
        f"- Needs more information: **{counts.get('Needs More Information', 0)}**",
        f"- Package/image automation candidates: **{len(result['automation_candidates'])}**",
        "",
        "## Prioritized Package/Image Candidates",
        "",
    ]
    if not result["automation_candidates"]:
        lines.append("No repeated package/image candidates were identified.")
    for index, candidate in enumerate(result["automation_candidates"][:max_candidates], start=1):
        lines.extend(
            [
                f"### {index}. {candidate['component']} {candidate['installed_version']}",
                "",
                f"- **Highest severity:** {candidate['severity']}",
                f"- **Distinct findings:** {candidate['finding_count']}",
                f"- **Affected resources:** {candidate['resource_count']}",
                f"- **Distribution/type:** {candidate['distro'] or 'Not provided'} / {candidate['resource_type']}",
                f"- **Different reported fix targets:** {candidate['fix_status_count']}",
                f"- **Example finding IDs:** {_examples(candidate['finding_ids'])}",
                f"- **Suggested route:** {candidate['suggested_route']}",
                "",
            ]
        )
    if len(result["automation_candidates"]) > max_candidates:
        lines.extend(
            [
                f"Showing the top {max_candidates} of {len(result['automation_candidates'])} candidates.",
                "",
            ]
        )
    lines.extend(["## Prioritized Finding Patterns", ""])
    if not result["patterns"]:
        lines.append("No complete patterns were identified.")
    for index, pattern in enumerate(result["patterns"][:max_patterns], start=1):
        component = f" / {pattern['component']}" if pattern["component"] else ""
        lines.extend(
            [
                f"### {index}. [{pattern['category']}] {pattern['finding_id']}{component}",
                "",
                f"- **Severity:** {pattern['severity']}",
                f"- **Affected resources:** {pattern['resource_count']}",
                f"- **Example resources:** {_examples(pattern['resources'])}",
                f"- **Reported fix status:** {pattern['fix_status'] or 'Not provided'}",
                f"- **Terraform status:** {pattern['management_status']}",
                f"- **Suggested route:** {pattern['suggested_route']}",
                "",
            ]
        )
    if len(result["patterns"]) > max_patterns:
        lines.extend(
            [
                f"Showing the top {max_patterns} of {len(result['patterns'])} patterns. "
                "Use the categorized CSV for the complete row-level result.",
                "",
            ]
        )
    lines.extend(
        [
            "## Human Review Gates",
            "",
            "- Confirm the resource/image owner and business context.",
            "- Verify whether the affected host is recreated from an image/template or changed in place.",
            "- Treat reported fix versions as evidence to review, not an automatically selected target.",
            "- Verify code and state before assigning Terraform ownership; absence of evidence means `Unknown`.",
            "- Validate compatibility, scope, rollout, dependencies, impact, rollback, and approval before action.",
            "",
        ]
    )
    return "\n".join(lines)


def write_categorized_csv(rows: list[dict[str, str]], path: Union[str, Path]) -> None:
    """Write normalized row-level classifications for human review."""
    fieldnames = list(DEFAULT_CONFIG["column_aliases"]) + ["category", "suggested_route"]
    with Path(path).open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def run_triage(
    input_path: Union[str, Path], config_path: Union[str, Path] = "config.yaml"
) -> dict[str, Any]:
    """Run Phase 1 triage and write the configured review artifacts."""
    config = load_config(config_path)
    rows = read_findings(input_path, sheet_name=config.get("sheet_name"))
    findings = normalize_findings(rows, config["column_aliases"])
    result = classify_findings(findings, int(config["recurring_min_resources"]))
    report = render_report(
        result,
        max_candidates=int(config.get("max_candidates_in_report", 25)),
        max_patterns=int(config.get("max_patterns_in_report", 50)),
    )
    Path(config["output"]["report_path"]).write_text(report, encoding="utf-8")
    write_categorized_csv(result["categorized_rows"], config["output"]["categorized_path"])
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Classify Prisma findings for human triage")
    parser.add_argument("input", help="Path to a Prisma CSV or XLSX report")
    parser.add_argument("--config", default="config.yaml", help="Path to YAML configuration")
    args = parser.parse_args()
    result = run_triage(args.input, args.config)
    print(f"Classified {result['total']} rows. No infrastructure changes were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
