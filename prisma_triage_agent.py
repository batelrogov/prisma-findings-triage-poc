"""Phase 1 Prisma findings triage.

This module classifies report rows into recurring patterns, unique findings,
or findings that need more information. It never changes cloud resources.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable, Optional, Union

import yaml


DEFAULT_CONFIG = {
    "recurring_min_count": 3,
    "sheet_name": None,
    "output": {
        "report_path": "triage_report.md",
        "categorized_path": "categorized_findings.csv",
    },
    "column_aliases": {
        "policy": ["policy", "policy name", "finding type", "alert name"],
        "resource_id": ["resource id", "resource", "resource name", "asset id"],
        "resource_type": ["resource type", "asset type", "cloud type"],
        "severity": ["severity", "risk", "priority"],
        "account": ["account", "account name", "account id", "cloud account"],
        "region": ["region", "cloud region"],
        "description": ["description", "finding description", "issue"],
        "remediation": ["remediation", "recommendation", "resolution"],
        "owner": ["owner", "resource owner", "team"],
        "terraform_status": [
            "terraform status",
            "terraform managed",
            "iac status",
            "management status",
        ],
    },
}

SEVERITY_WEIGHT = {
    "critical": 5,
    "high": 4,
    "medium": 3,
    "low": 2,
    "informational": 1,
    "info": 1,
}


def load_config(config_path: Union[str, Path] = "config.yaml") -> dict[str, Any]:
    """Load YAML configuration and merge it over safe defaults."""
    config = {
        **DEFAULT_CONFIG,
        "output": {**DEFAULT_CONFIG["output"]},
        "column_aliases": {
            key: list(values)
            for key, values in DEFAULT_CONFIG["column_aliases"].items()
        },
    }
    path = Path(config_path)
    if not path.exists():
        return config

    loaded = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    for key, value in loaded.items():
        if key in {"output", "column_aliases"} and isinstance(value, dict):
            config[key].update(value)
        else:
            config[key] = value
    return config


def _normalized_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_xlsx(path: Path, sheet_name: Optional[str]) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on local installation
        raise RuntimeError("XLSX input requires openpyxl from requirements.txt") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [_clean(value) for value in next(rows)]
    except StopIteration:
        return []
    return [dict(zip(headers, row)) for row in rows]


def read_findings(
    path: Union[str, Path], sheet_name: Optional[str] = None
) -> list[dict[str, Any]]:
    """Read Prisma findings from CSV or XLSX."""
    input_path = Path(path)
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(input_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(input_path, sheet_name)
    raise ValueError("Unsupported input format. Use CSV or XLSX.")


def normalize_findings(
    rows: Iterable[dict[str, Any]], column_aliases: dict[str, list[str]]
) -> list[dict[str, str]]:
    """Map report-specific column names to the canonical triage fields."""
    normalized = []
    alias_map = {
        field: {_normalized_header(alias) for alias in aliases + [field]}
        for field, aliases in column_aliases.items()
    }

    for source_row in rows:
        source = {_normalized_header(key): _clean(value) for key, value in source_row.items()}
        item = {
            field: next(
                (source[alias] for alias in aliases if source.get(alias)), ""
            )
            for field, aliases in alias_map.items()
        }
        normalized.append(item)
    return normalized


def _pattern_key(finding: dict[str, str]) -> tuple[str, str, str] | None:
    policy = _normalized_header(finding["policy"])
    resource_type = _normalized_header(finding["resource_type"])
    remediation = _normalized_header(finding["remediation"])
    if not policy or not resource_type:
        return None
    return policy, resource_type, remediation


def _reported_management_status(values: Iterable[str]) -> str:
    normalized = {_normalized_header(value) for value in values if value}
    managed = {"yes", "true", "managed", "terraform", "terraform managed"}
    unmanaged = {"no", "false", "unmanaged", "manual", "manually managed"}
    if normalized and normalized <= managed:
        return "Reported as Terraform-managed — validation required"
    if normalized and normalized <= unmanaged:
        return "Reported as unmanaged — validation required"
    if normalized:
        return "Mixed or unrecognized — validation required"
    return "Unknown — ownership validation required"


def _route(category: str, management_status: str) -> str:
    if category == "Needs More Information":
        return "Collect missing policy/resource/ownership context before routing"
    if category == "Unique":
        return "Manual review with the resource owner"
    if management_status.startswith("Reported as Terraform-managed"):
        return "Terraform template candidate after code/state ownership validation"
    if management_status.startswith("Reported as unmanaged"):
        return "Script-assisted or import candidate after human review"
    return "Validate ownership, then assess Terraform template or script-assisted treatment"


def classify_findings(
    findings: list[dict[str, str]], recurring_min_count: int = 3
) -> dict[str, Any]:
    """Classify normalized findings without making infrastructure changes."""
    if recurring_min_count < 2:
        raise ValueError("recurring_min_count must be at least 2")

    groups: dict[tuple[str, str, str], list[dict[str, str]]] = defaultdict(list)
    incomplete: list[dict[str, str]] = []
    for finding in findings:
        key = _pattern_key(finding)
        if key is None or not finding["resource_id"]:
            incomplete.append(finding)
        else:
            groups[key].append(finding)

    patterns = []
    categorized_rows = []
    for key, members in groups.items():
        category = "Recurring" if len(members) >= recurring_min_count else "Unique"
        management_status = _reported_management_status(
            member["terraform_status"] for member in members
        )
        severity = max(
            (member["severity"] for member in members),
            key=lambda value: SEVERITY_WEIGHT.get(value.lower(), 0),
            default="Unknown",
        ) or "Unknown"
        pattern = {
            "key": key,
            "category": category,
            "policy": members[0]["policy"],
            "resource_type": members[0]["resource_type"],
            "severity": severity,
            "count": len(members),
            "accounts": sorted({m["account"] for m in members if m["account"]}),
            "resources": [m["resource_id"] for m in members],
            "management_status": management_status,
            "suggested_route": _route(category, management_status),
            "priority_score": SEVERITY_WEIGHT.get(severity.lower(), 0) * 100 + len(members),
        }
        patterns.append(pattern)
        for member in members:
            categorized_rows.append(
                {**member, "category": category, "suggested_route": pattern["suggested_route"]}
            )

    for finding in incomplete:
        categorized_rows.append(
            {
                **finding,
                "category": "Needs More Information",
                "suggested_route": _route("Needs More Information", "Unknown"),
            }
        )

    patterns.sort(key=lambda item: item["priority_score"], reverse=True)
    counts = Counter(row["category"] for row in categorized_rows)
    return {
        "total": len(findings),
        "counts": dict(counts),
        "patterns": patterns,
        "categorized_rows": categorized_rows,
    }


def render_report(result: dict[str, Any]) -> str:
    """Render a review-oriented Markdown report."""
    counts = result["counts"]
    lines = [
        "# Prisma Findings Triage Report",
        "",
        "> Classification draft only. No cloud resources, scripts, or Terraform changes were executed.",
        "",
        "## Executive Summary",
        "",
        f"- Total findings: **{result['total']}**",
        f"- Recurring findings: **{counts.get('Recurring', 0)}**",
        f"- Unique findings: **{counts.get('Unique', 0)}**",
        f"- Needs more information: **{counts.get('Needs More Information', 0)}**",
        "",
        "## Prioritized Patterns",
        "",
    ]
    if not result["patterns"]:
        lines.append("No complete patterns were identified.")
    for index, pattern in enumerate(result["patterns"], start=1):
        resources = ", ".join(f"`{value}`" for value in pattern["resources"][:5])
        if len(pattern["resources"]) > 5:
            resources += f" and {len(pattern['resources']) - 5} more"
        lines.extend(
            [
                f"### {index}. [{pattern['category']}] {pattern['policy']}",
                "",
                f"- **Severity:** {pattern['severity']}",
                f"- **Resource type:** {pattern['resource_type']}",
                f"- **Finding count:** {pattern['count']}",
                f"- **Example resources:** {resources}",
                f"- **Terraform status:** {pattern['management_status']}",
                f"- **Suggested route:** {pattern['suggested_route']}",
                "",
            ]
        )
    lines.extend(
        [
            "## Human Review Gates",
            "",
            "- Confirm the resource owner and business context.",
            "- Verify Terraform ownership using both code and state; absence of evidence means `Unknown`, not `Manual`.",
            "- Validate scope, dependencies, impact, rollback, and approval before any action.",
            "- Treat script-assisted and Terraform routes as candidates, not approved remediation.",
            "",
        ]
    )
    return "\n".join(lines)


def write_categorized_csv(
    rows: list[dict[str, str]], path: Union[str, Path]
) -> None:
    """Write row-level classifications for filtering and human review."""
    output_path = Path(path)
    fieldnames = list(DEFAULT_CONFIG["column_aliases"]) + ["category", "suggested_route"]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def run_triage(
    input_path: Union[str, Path],
    config_path: Union[str, Path] = "config.yaml",
) -> dict[str, Any]:
    """Run Phase 1 triage and write the configured review artifacts."""
    config = load_config(config_path)
    rows = read_findings(input_path, sheet_name=config.get("sheet_name"))
    findings = normalize_findings(rows, config["column_aliases"])
    result = classify_findings(findings, int(config["recurring_min_count"]))
    Path(config["output"]["report_path"]).write_text(
        render_report(result), encoding="utf-8"
    )
    write_categorized_csv(result["categorized_rows"], config["output"]["categorized_path"])
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Classify Prisma findings for human triage")
    parser.add_argument("input", help="Path to a Prisma CSV or XLSX report")
    parser.add_argument("--config", default="config.yaml", help="Path to YAML configuration")
    args = parser.parse_args()
    result = run_triage(args.input, args.config)
    print(f"Classified {result['total']} findings. No infrastructure changes were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
